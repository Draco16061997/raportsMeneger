import os
import getDb
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

import paths
import users



path = paths.save_exel # путь сохранение отчетов




def createRaport(start, end):
    file_name = f"Звіт від {start} до {end}.xlsx"
    file_path = path + file_name

    if not os.path.exists(file_path):
        wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
        sheet_title1 = f" общий отчёт {start} до {end} "
        wb.create_sheet(index=0, title=sheet_title1)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title1]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteDB(False, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h


        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteDB(True, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h

# =============================================

        sheet_title2 = f" ДО отчёт {start} до {end} "
        wb.create_sheet(index=1, title=sheet_title2)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title2]



        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteResDB(False,0,start,end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteResDB(True,0,start,end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h


# ===================================================




        sheet_title4 = f" НМ отчёт {start} до {end} "
        wb.create_sheet(index=2, title=sheet_title4)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title4]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteResDB(False, 1, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteResDB(True, 1, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h

# =================================================



        sheet_title3 = f" ДП отчёт {start} до {end} "
        wb.create_sheet(index=3, title=sheet_title3)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title3]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteResDB(False, 2, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteResDB(True, 2, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h

        # =================================================

        sheet_title01 = f" Дарина Сухоніс  отчёт {start} до {end} "
        wb.create_sheet(index=4, title=sheet_title01)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title01]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(False,1,start,end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(True,1, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h

        # =================================================

        sheet_title02 = f" Анастасія Ткач отчёт {start} до {end} "
        wb.create_sheet(index=5, title=sheet_title02)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title02]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(False, 2, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(True, 2, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h

        # =================================================

        sheet_title03 = f" Валерія Малихіна отчёт {start} до {end} "
        wb.create_sheet(index=6, title=sheet_title03)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title03]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(False, 3, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(True, 3, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h

        # =================================================

        sheet_title04 = f" Андрій Нікітін отчёт {start} до {end} "
        wb.create_sheet(index=7, title=sheet_title04)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title04]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(False, 4, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(True, 4, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h

        # =================================================

        sheet_title05 = f" Антон Глущенко отчёт {start} до {end} "
        wb.create_sheet(index=8, title=sheet_title05)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title05]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20
        s = 1

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(False, 5, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        sheet.append((" ", " "))
        sheet.append((" ", " "))
        sheet.row_dimensions[n + 1].height = h
        sheet.append(("Дата публікації ", "SHORTS", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n + 1].height = h

        sheet.row_dimensions[n].height = h
        for i in getDb.ExecuteJurnDB(True, 5, start, end):
            sheet.append(i)
            s += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО Shorts", s - 1))
        sheet.row_dimensions[n + 1].height = h




        wb.save(file_path)
#
# for i in getDb.ExecuteResDB(False,0,'2023-11-01','2023-11-04'):
#     print(i)

if __name__ == "__main__":
    createRaport('2023-11-01', '2023-11-04')
    print(f"отчёт лежит {path}")



    # CreateExel.createReportIsChanel('Дніпро Оперативний - життя великого міста', d, period_to_now(30))
    # CreateExel.createReportIsChanel('Наше Місто', d, period_to_now(30))
    # CreateExel.createReportIsChanel('Дніпровська Панорама', d, period_to_now(30))

    # с.createReportIsTimeAndName("Малихіна", d, period_to_now(30))
    # с.createReportIsTimeAndName('Ткач', d, period_to_now(30))
    # с.createReportIsTimeAndName('Глущенко', d, period_to_now(30))
    # с.createReportIsTimeAndName('Сухоніс', d, period_to_now(30))
    # с.createReportIsTimeAndName('Нікітін', d, period_to_now(30))

    # с.createReportIsTimeAndName('Скрипніченко', d, period_to_now(30))
    # с.createReportIsTimeAndName('Бездільний', d, period_to_now(30))
    # с.createReportIsTimeAndName('Тарасов', d, period_to_now(30))
    # с.createReportIsTimeAndName('Тейлор', d, period_to_now(30))
    # с.createReportIsTimeAndName('Нікітін', d, period_to_now(30))

    # с.createReportIsTime(d, period_to_now(30))
