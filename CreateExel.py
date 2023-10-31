import os
import getIsDatabase
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook




path = "/volume2/Montage/Users/reports/" # путь сохранение отчетов

def createReportIsTime(start, end):
    file_name = f"Звіт по сюжетам за период від {start} до {end}.xlsx"
    file_path = path + file_name

    if not os.path.exists(file_path):
        wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
        sheet_title = f" отчёт {start} до {end}"
        wb.create_sheet(index=0, title=sheet_title)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getIsDatabase.getPeriod(start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        wb.save(file_path)

def createReportIsTimeAndName(name, start, end):
    file_name = f"Звіт від {start} до {end} {name}.xlsx"
    file_path = path + file_name

    if not os.path.exists(file_path):
        wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
        sheet_title = f" отчёт {start} до {end}"
        wb.create_sheet(index=0, title=sheet_title)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getIsDatabase.getJurnForPeriod(name, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        wb.save(file_path)

def createReportIsChanel(name, start, end):
    file_name = f"Звіт від {start} до {end} канал: {name}.xlsx"
    file_path = path + file_name

    if not os.path.exists(file_path):
        wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
        sheet_title = f" отчёт {start} до {end}"
        wb.create_sheet(index=0, title=sheet_title)
        try:
            wb.remove(wb["Sheet"])
        except:
            pass

        sheet = wb[sheet_title]

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 100
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 50
        n = 1
        h = 20

        sheet.append(("Дата публікації ", "Тип сюжету", "Назва ", "ЮТУБ КАНАЛ", "Журналіст ", "Монтував", "Url"))
        sheet.row_dimensions[n].height = h
        for i in getIsDatabase.getRecurceList(name, start, end):
            sheet.append(i)
            n += 1
            sheet.row_dimensions[n].height = h
        sheet.append(("ВСЬОГО СЮЖЕТІВ", n - 1))
        sheet.row_dimensions[n + 1].height = h

        wb.save(file_path)






if __name__ == "__main__":
    print(g.getNameDir())
    print(g.getNameJurn())
    print(g.getResurce())


    # CreateExel.createReportIsChanel('Дніпро Оперативний - життя великого міста', d, period_to_now(30))
    # CreateExel.createReportIsChanel('Наше Місто', d, period_to_now(30))
    # CreateExel.createReportIsChanel('Дніпровська Панорама', d, period_to_now(30))

    # с.createReportIsTimeAndName("Малихіна", d, period_to_now(30))
    # с.createReportIsTimeAndName('Ткач', d, period_to_now(30))
    # с.createReportIsTimeAndName('Глущенко', d, period_to_now(30))
    # с.createReportIsTimeAndName('Сухоніс', d, period_to_now(30))
    # с.createReportIsTimeAndName('Нікітін', d, period_to_now(30))

    # с.createReportIsTimeAndName('Скрипніченко', d, period_to_now(30))
    # с.createReportIsTimeAndName('Бездільний', d, period_to_now(30))
    # с.createReportIsTimeAndName('Тарасов', d, period_to_now(30))
    # с.createReportIsTimeAndName('Тейлор', d, period_to_now(30))
    # с.createReportIsTimeAndName('Нікітін', d, period_to_now(30))

    # с.createReportIsTime(d, period_to_now(30))
